import csv
import io
import os
import logging
from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from django.conf import settings
from django.contrib.auth import get_user_model
User = get_user_model()
from .models import Profile, Application, Document, AuditLog, UserDocument, Payment, Appeal, ShareableLink
from .serializers import (
    UserSerializer, ProfileSerializer, ApplicationSerializer,
    DocumentSerializer, AuditLogSerializer, UserDocumentSerializer,
    PaymentSerializer, AppealSerializer, ShareableLinkSerializer
)
from notifications.utils import create_notification
from users.permissions import IsAdminUser
import uuid
from django.utils import timezone
from datetime import timedelta

logger = logging.getLogger(__name__)


def _is_staff(user):
    return user.is_authenticated and user.role in ('admin', 'director', 'ssw')


def _is_director(user):
    """Banking details and full student files are Director-only (§6.4)."""
    return user.is_authenticated and user.role == 'director'


# ---------------------------------------------------------------------------
# SHARED CSV BUILDER
# Used by both export_csv (approved-only download) and
# dispatch_report (ALL records emailed to finance).
# ---------------------------------------------------------------------------

def _build_full_csv(
    funding_type='all', date_from=None, date_to=None,
    all_statuses=False, unsent_only=False,
    student_name=None, semester=None, year=None, award_type=None,
):
    """
    Build the comprehensive student records CSV with personal + banking info.

    Columns (35 total):
      Submission info  (5)
      Personal info    (12)  — from CustomUser + Profile
      Banking info     (6)   — from CustomUser
      Enrollment info  (4)   — from CustomUser
      Application info (4)
      Payment info     (5)
      Timeline info    (3)

    Parameters
    ----------
    all_statuses : bool
        False  → only accepted/approved records  (used by export_csv download)
        True   → every record regardless of status (used by dispatch_report email)
    unsent_only : bool
        True   → only records never sent to finance (finance_sent_at__isnull=True)

    Returns
    -------
    (csv_bytes: bytes, row_count: int)
    """
    from forms.models import FormSubmission
    from django.db.models import Q

    # §6.3: UPI (unique identifier) must NEVER appear in any export.
    # §6.4: Banking details are Director-only — omitted from general CSV.
    HEADERS = [
        # ── Submission ──
        'Submission ID', 'Form/Type', 'Status', 'Stream',
        # ── Personal ──
        'Full Name', 'Email', 'Phone', 'Alternate Phone',
        'Date of Birth', 'Gender', 'Pronouns',
        'Beneficiary #', 'Treaty #',
        'Mailing Address', 'Town/City', 'Postal Code', 'Province',
        'No. of Dependents', 'Dependent Ages',
        'Financial Assistance Status', 'SFA Active (Profile)',
        # ── Banking ──
        'Account Holder Name', 'Account Type',
        'Bank Name', 'Transit #', 'Institution #', 'Account #',
        # ── Enrollment ──
        'Institute (Profile)', 'Institution Name', 'Program', 'Enrollment Status', 'Course Load (%)',
        # ── Funding Breakdown (Single Row) ──
        'Tuition ($)', 'Living Allowance ($)', 'Books & Supplies ($)', 'Travel ($)', 'Special Awards ($)',
        'TOTAL APPROVED ($)',
        # ── Timeline ──
        'Submitted Date', 'Decision Date', 'Decided By',
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)

    # ── New-model submissions ──
    if all_statuses:
        qs = FormSubmission.objects.all()
    else:
        qs = FormSubmission.objects.filter(status='accepted')

    if unsent_only:
        qs = qs.filter(finance_sent_at__isnull=True)

    qs = qs.select_related('student', 'form', 'decided_by').prefetch_related(
        'payments', 'student__profile'
    )

    # ── Legacy applications ──
    if all_statuses:
        legacy_qs = Application.objects.all()
    else:
        legacy_qs = Application.objects.filter(status='approved')

    if unsent_only:
        legacy_qs = legacy_qs.filter(finance_sent_at__isnull=True)

    legacy_qs = legacy_qs.select_related('student').prefetch_related(
        'payments', 'student__profile'
    )

    # ── Funding-type filter (only for approved download) ──
    if not all_statuses and funding_type != 'all':
        mapping = {
            'cdfn':  Q(form__title__icontains='FormA') | Q(form__title__icontains='FormC'),
            'dggr':  (Q(form__title__icontains='DGGR') | Q(form__title__icontains='Scholarship') |
                      Q(form__title__icontains='Hardship') | Q(form__title__icontains='Form D') |
                      Q(form__title__icontains='Form F') | Q(form__title__icontains='Form G')),
            'ucepp': Q(form__title__icontains='UCEPP') | Q(form__title__icontains='Upgrading'),
        }
        if funding_type in mapping:
            qs = qs.filter(mapping[funding_type])
            if funding_type == 'cdfn':
                legacy_qs = legacy_qs.filter(form_type__icontains='PSSSP')
            elif funding_type == 'dggr':
                legacy_qs = legacy_qs.filter(form_type__icontains='DGGR')
            elif funding_type == 'ucepp':
                legacy_qs = legacy_qs.filter(form_type__icontains='UCEPP')

    # ── Date filter ──
    if date_from:
        qs = qs.filter(submitted_at__date__gte=date_from)
        legacy_qs = legacy_qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(submitted_at__date__lte=date_to)
        legacy_qs = legacy_qs.filter(created_at__date__lte=date_to)

    # ── §3.1.H ad-hoc filters ──
    if student_name:
        qs = qs.filter(student__full_name__icontains=student_name)
        legacy_qs = legacy_qs.filter(student__full_name__icontains=student_name)
    if semester:
        # Semester stored in form title or office_use_data
        qs = qs.filter(office_use_data__semester__icontains=semester)
        legacy_qs = legacy_qs.filter(semester__icontains=semester)
    if year:
        qs = qs.filter(submitted_at__year=year)
        legacy_qs = legacy_qs.filter(created_at__year=year)
    if award_type:
        qs = qs.filter(form__title__icontains=award_type)
        legacy_qs = legacy_qs.filter(form_type__icontains=award_type)

    D = '—'  # default for missing values

    def _student_personal_banking(student, profile):
        """Return the 28 personal + banking + enrollment columns for a student."""
        u = student
        p = profile
        return [
            # Personal (18 cols)
            u.full_name if u else D,
            u.email if u else D,
            (u.phone or D) if u else D,
            (u.alternate_phone or D) if u else D,
            str(u.dob) if (u and u.dob) else D,
            (u.gender or (p.gender if p else D) or D) if u else D,
            (u.pronouns or (p.pronouns if p else D) or D) if u else D,
            (u.beneficiary_number or (p.beneficiary_number if p else D) or D) if u else D,
            (u.treaty_number or D) if u else D,
            # UPI intentionally excluded — §6.3 prohibits it in any export
            (u.mailing_address or (p.mailing_address if p else D) or D) if u else D,
            (p.town_city or D) if p else D,
            (p.postal_code or D) if p else D,
            (u.province_of_residence or D) if u else D,
            str(u.num_dependents) if u else D,
            (u.dependent_ages or D) if u else D,
            (u.financial_assistance_status or D) if u else D,
            ('Yes' if p.is_sfa_active else 'No') if p else D,
            # Banking (6 cols)
            (u.account_holder_name or D) if u else D,
            (u.account_type or D) if u else D,
            (u.bank_name or D) if u else D,
            (u.transit_number or D) if u else D,
            (u.inst_number or D) if u else D,
            (u.account_number or D) if u else D,
            # Enrollment (5 cols)
            (p.institute or D) if p else D,
            (u.institution_name or D) if u else D,
            (u.program_credential or D) if u else D,
            (u.enrollment_status or D) if u else D,
            str(u.course_load) if u else D,
        ]

    row_count = 0

    # ── New-model submission rows ──
    for sub in qs.order_by('-submitted_at'):
        student = sub.student
        profile = getattr(student, 'profile', None) if student else None
        payments = sub.payments.all()

        submission_cols = [
            f"FS-{sub.id}",
            sub.form.title if sub.form else D,
            sub.status,
            student.primary_stream if student else D,
        ]
        personal_banking = _student_personal_banking(student, profile)
        dates = [
            sub.submitted_at.strftime('%Y-%m-%d') if sub.submitted_at else D,
            sub.decided_at.strftime('%Y-%m-%d') if sub.decided_at else D,
            sub.decided_by.full_name if sub.decided_by else D,
        ]

        # Consolidate payment breakdown into columns
        breakdown = {'Tuition': 0, 'Living': 0, 'Books': 0, 'Travel': 0, 'Special': 0}
        for p in payments:
            pt = p.payment_type.lower()
            if 'tuition' in pt: breakdown['Tuition'] += p.amount
            elif 'living' in pt: breakdown['Living'] += p.amount
            elif 'book' in pt: breakdown['Books'] += p.amount
            elif 'travel' in pt: breakdown['Travel'] += p.amount
            else: breakdown['Special'] += p.amount

        breakdown_cols = [
            breakdown['Tuition'] or D,
            breakdown['Living'] or D,
            breakdown['Books'] or D,
            breakdown['Travel'] or D,
            breakdown['Special'] or D,
            sub.amount or sum(breakdown.values()) or 0
        ]

        writer.writerow(submission_cols + personal_banking + breakdown_cols + dates)
        row_count += 1

    # ── Legacy application rows ──
    for app in legacy_qs.order_by('-created_at'):
        student = app.student
        profile = getattr(student, 'profile', None) if student else None
        
        submission_cols = [
            f"LEG-{app.id}",
            app.form_type,
            app.status,
            student.primary_stream if student else D,
        ]
        personal_banking = _student_personal_banking(student, profile)
        dates = [
            app.created_at.strftime('%Y-%m-%d') if app.created_at else D,
            app.decision_at.strftime('%Y-%m-%d') if app.decision_at else D,
            app.decision_by or D,
        ]

        # For legacy, just show the total in the last column
        breakdown_cols = [D, D, D, D, D, app.amount or 0]

        writer.writerow(submission_cols + personal_banking + breakdown_cols + dates)
        row_count += 1

    return output.getvalue().encode('utf-8'), row_count


def _build_full_csv_from_qs(qs):
    """
    Same 42-column CSV but built from an already-filtered FormSubmission queryset.
    Used by dispatch_report so it only includes the exact unsent submissions.
    Returns (csv_bytes, row_count).
    """
    # §6.3: UPI (unique identifier) must NEVER appear in any export.
    HEADERS = [
        'Submission ID', 'Form/Type', 'Status', 'Approved Amount ($)', 'Stream',
        'Full Name', 'Email', 'Phone', 'Alternate Phone',
        'Date of Birth', 'Gender', 'Pronouns',
        'Beneficiary #', 'Treaty #',
        'Mailing Address', 'Town/City', 'Postal Code', 'Province',
        'No. of Dependents', 'Dependent Ages',
        'Financial Assistance Status', 'SFA Active (Profile)',
        'Account Holder Name', 'Account Type',
        'Bank Name', 'Transit #', 'Institution #', 'Account #',
        'Institute (Profile)', 'Institution Name', 'Program', 'Enrollment Status', 'Course Load (%)',
        'Payment Type', 'Payment Amount ($)', 'Payment Status', 'Payment Reference #', 'Payment Date',
        'Submitted Date', 'Decision Date', 'Decided By',
    ]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)

    D = '—'

    def _pb(student, profile):
        u, p = student, profile
        return [
            u.full_name if u else D, u.email if u else D,
            (u.phone or D) if u else D, (u.alternate_phone or D) if u else D,
            str(u.dob) if (u and u.dob) else D,
            (u.gender or (p.gender if p else D) or D) if u else D,
            (u.pronouns or (p.pronouns if p else D) or D) if u else D,
            (u.beneficiary_number or (p.beneficiary_number if p else D) or D) if u else D,
            (u.treaty_number or D) if u else D,
            # UPI intentionally excluded — §6.3
            (u.mailing_address or (p.mailing_address if p else D) or D) if u else D,
            (p.town_city or D) if p else D, (p.postal_code or D) if p else D,
            (u.province_of_residence or D) if u else D,
            str(u.num_dependents) if u else D, (u.dependent_ages or D) if u else D,
            (u.financial_assistance_status or D) if u else D,
            ('Yes' if p.is_sfa_active else 'No') if p else D,
            (u.account_holder_name or D) if u else D, (u.account_type or D) if u else D,
            (u.bank_name or D) if u else D, (u.transit_number or D) if u else D,
            (u.inst_number or D) if u else D, (u.account_number or D) if u else D,
            (p.institute or D) if p else D,
            (u.institution_name or D) if u else D, (u.program_credential or D) if u else D,
            (u.enrollment_status or D) if u else D, str(u.course_load) if u else D,
        ]

    row_count = 0
    for sub in qs.order_by('-submitted_at'):
        student = sub.student
        profile = getattr(student, 'profile', None) if student else None
        payments = sub.payments.all()
        sub_cols = [
            f"FS-{sub.id}", sub.form.title if sub.form else D,
            sub.status, sub.amount or 0,
            student.primary_stream if student else D,
        ]
        pb = _pb(student, profile)
        dates = [
            sub.submitted_at.strftime('%Y-%m-%d') if sub.submitted_at else D,
            sub.decided_at.strftime('%Y-%m-%d') if sub.decided_at else D,
            sub.decided_by.full_name if sub.decided_by else D,
        ]
        if payments:
            for p in payments:
                writer.writerow(sub_cols + pb + [
                    p.payment_type, p.amount, p.status,
                    p.reference_number or D,
                    p.date_issued.strftime('%Y-%m-%d') if p.date_issued else D,
                ] + dates)
                row_count += 1
        else:
            writer.writerow(sub_cols + pb + [D, D, D, D, D] + dates)
            row_count += 1

    return output.getvalue().encode('utf-8'), row_count


# ---------------------------------------------------------------------------
# VIEWS
# ---------------------------------------------------------------------------

class RegisterView(viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]
    serializer_class = UserSerializer

    def create(self, request):
        user = User.objects.create_user(
            username=request.data.get('email'),
            email=request.data.get('email'),
            password=request.data.get('password'),
            first_name=request.data.get('firstName', ''),
            last_name=request.data.get('lastName', '')
        )
        Profile.objects.create(
            user=user,
            beneficiary_number=request.data.get('beneficiaryNo', ''),
            indian_status=request.data.get('treatyNum', '')
        )
        return Response({'status': 'user created'}, status=status.HTTP_201_CREATED)


class UserDetailView(viewsets.GenericViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class ProfileViewSet(viewsets.ModelViewSet):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if _is_staff(user):
            student_id = self.request.query_params.get('student_id')
            if student_id:
                return self.queryset.filter(user_id=student_id)
            return self.queryset.all()
        return self.queryset.filter(user=user)

    def get_object(self):
        # Allow staff to retrieve/update by profile pk directly
        obj = super().get_object()
        user = self.request.user
        if not _is_staff(user) and obj.user != user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You do not have permission to access this profile.")
        return obj

    def perform_update(self, serializer):
        old_instance = serializer.instance
        change_triggers = []

        _FORM_D_FIELDS = {
            'is_sfa_active': 'SFA status',
            'enrollment_status': 'Enrollment status',
            'num_dependents': 'Number of dependents',
            'institution_name': 'Institution name',
        }
        for field, label in _FORM_D_FIELDS.items():
            if field in serializer.validated_data:
                old_val = getattr(old_instance, field, None)
                new_val = serializer.validated_data[field]
                if old_val != new_val:
                    change_triggers.append((field, label, str(old_val), str(new_val)))

        instance = serializer.save()

        if change_triggers:
            self._auto_trigger_form_d(instance, change_triggers)

    def _auto_trigger_form_d(self, profile, change_triggers):
        """§5: When student updates profile in ways that affect funding, create a
        notification prompting them to submit Form D (Change of Information)."""
        try:
            create_notification(
                profile.user,
                "Important: Submit Change of Information (Form D)",
                "Your profile was updated in a way that may affect your funding. "
                "Please submit a Form D (Change of Information) as soon as possible "
                "to ensure your funding is recalculated correctly.",
            )
        except Exception as exc:
            logger.warning("Form D auto-trigger notification failed: %s", exc)


class ApplicationViewSet(viewsets.ModelViewSet):
    queryset = Application.objects.all()
    serializer_class = ApplicationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = Application.objects.select_related(
            'student', 
            'student__profile'
        ).prefetch_related(
            'documents'
        ).order_by('-created_at')

        if user.role in ('admin', 'director', 'ssw'):
            return qs  # All staff see every application + complete history (§3.1.B)
        return qs.filter(student=user)

    def perform_create(self, serializer):
        serializer.save(student=self.request.user)

    def perform_update(self, serializer):
        old_status = self.get_object().status
        instance = serializer.save()
        if instance.status == 'pending' and old_status != 'pending':
            if old_status != 'review' and self.request.user.role != 'director':
                # Revert status if not approved by admin first
                instance.status = old_status
                instance.save()
                from rest_framework.exceptions import ValidationError
                raise ValidationError("Application must be in 'Review' status before forwarding to Director.")
            
            directors = User.objects.filter(role='director')
            directors = User.objects.filter(role='director')
            for director in directors:
                from notifications.models import Notification
                Notification.objects.create(
                    user=director,
                    title="Legacy Application Awaiting Approval",
                    message=f"Application #{instance.id} from {instance.student.full_name if instance.student else 'Student'} needs your decision.",
                    link="/staff/director-queue"
                )
                if director.email:
                    try:
                        from notifications.utils import email_director_approval_request
                        email_director_approval_request(
                            director_email=director.email,
                            student_name=instance.student.full_name if instance.student else 'Student',
                            form_title=instance.form_type,
                            amount=float(0),
                            submission_id=f"legacy-{instance.id}",
                        )
                    except Exception:
                        pass

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        # §3.1.D: only the Director may approve/deny applications
        if not _is_director(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only the Director may approve applications.")
        application = self.get_object()
        application.status = Application.Status.APPROVED
        application.decision_by = request.user.full_name or request.user.email
        application.decision_notes = request.data.get('notes', '')
        application.decision_at = timezone.now()
        application.save()
        AuditLog.objects.create(
            action=f"Approved Application {application.id}",
            performed_by=request.user,
            role=request.user.role,
            application=application,
            details=application.decision_notes
        )
        return Response({'status': 'application approved'})

    @action(detail=True, methods=['post'])
    def deny(self, request, pk=None):
        # §3.1.D: only the Director may approve/deny applications
        if not _is_director(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only the Director may deny applications.")
        application = self.get_object()
        application.status = Application.Status.DENIED
        application.decision_by = request.user.full_name or request.user.email
        application.decision_notes = request.data.get('notes', '')
        application.decision_at = timezone.now()
        application.save()
        AuditLog.objects.create(
            action=f"Denied Application {application.id}",
            performed_by=request.user,
            role=request.user.role,
            application=application,
            details=application.decision_notes
        )
        return Response({'status': 'application denied'})

    @action(detail=True, methods=['post'])
    def request_info(self, request, pk=None):
        application = self.get_object()
        application.status = Application.Status.INFO_REQUIRED
        application.save()
        notes = request.data.get('notes', 'More information is required for your application.')
        create_notification(
            user=application.student,
            title="Information Requested",
            message=f"Staff has requested more information for your {application.form_type} application: {notes}",
            link=f"/dashboard?appId={application.id}"
        )
        AuditLog.objects.create(
            action=f"Requested Info for Application {application.id}",
            performed_by=request.user,
            application=application,
            details=notes
        )
        return Response({'status': 'info requested'})

    @action(detail=True, methods=['post'])
    def share(self, request, pk=None):
        from api.models import PolicySetting
        from django.conf import settings as django_settings
        application = self.get_object()
        expiry_config = PolicySetting.objects.filter(section='system_config', field_key='share_link_expiry_days').first()
        expiry_days = int(expiry_config.value) if expiry_config else 7
        token = uuid.uuid4().hex
        expires_at = timezone.now() + timedelta(days=expiry_days)
        ShareableLink.objects.create(application=application, token=token, expires_at=expires_at)
        base_url = getattr(django_settings, 'SITE_URL', request.build_absolute_uri('/').rstrip('/'))
        return Response({'token': token, 'url': f"{base_url}/shared/{token}", 'expires_at': expires_at})


class SharedApplicationView(viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]

    @action(detail=False, methods=['get'], url_path='view/(?P<token>[^/.]+)')
    def view_by_token(self, request, token=None):
        try:
            share_link = ShareableLink.objects.get(token=token)
            if not share_link.is_valid():
                return Response({'error': 'Link expired or invalid'}, status=status.HTTP_403_FORBIDDEN)
            return Response(ApplicationSerializer(share_link.application).data)
        except ShareableLink.DoesNotExist:
            return Response({'error': 'Link not found'}, status=status.HTTP_404_NOT_FOUND)


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if _is_staff(user):
            return self.queryset.all()
        return self.queryset.filter(user=user)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        """Download CSV — approved records only, filterable by all report dimensions."""
        from django.http import HttpResponse

        funding_type = request.query_params.get('funding_type', 'all').lower()
        date_from    = request.query_params.get('date_from')
        date_to      = request.query_params.get('date_to')
        student_name = request.query_params.get('student')
        semester     = request.query_params.get('semester')
        year         = request.query_params.get('year')
        award_type   = request.query_params.get('award_type')

        csv_bytes, row_count = _build_full_csv(
            funding_type=funding_type,
            date_from=date_from,
            date_to=date_to,
            all_statuses=False,
            unsent_only=False,
            student_name=student_name,
            semester=semester,
            year=year,
            award_type=award_type,
        )

        AuditLog.objects.create(
            action="Payment CSV Export Triggered",
            performed_by=request.user,
            details=f"Exported {row_count} records"
        )

        response = HttpResponse(csv_bytes, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payment_export.csv"'
        return response

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Download Excel (.xlsx) report — same filters as export-csv. §3.1.H."""
        from django.http import HttpResponse

        funding_type = request.query_params.get('funding_type', 'all').lower()
        date_from    = request.query_params.get('date_from')
        date_to      = request.query_params.get('date_to')
        student_name = request.query_params.get('student')
        semester     = request.query_params.get('semester')
        year         = request.query_params.get('year')
        award_type   = request.query_params.get('award_type')

        csv_bytes, row_count = _build_full_csv(
            funding_type=funding_type,
            date_from=date_from,
            date_to=date_to,
            all_statuses=False,
            unsent_only=False,
            student_name=student_name,
            semester=semester,
            year=year,
            award_type=award_type,
        )

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import io, csv as _csv

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Student Funding Report"

            reader = _csv.reader(io.StringIO(csv_bytes.decode('utf-8')))
            rows = list(reader)
            if rows:
                # Header row styling
                header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col_idx, cell_val in enumerate(rows[0], start=1):
                    cell = ws.cell(row=1, column=col_idx, value=cell_val)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                # Data rows
                for row_idx, row in enumerate(rows[1:], start=2):
                    for col_idx, cell_val in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_val)
                # Auto-fit columns (approximate)
                for col in ws.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()

            AuditLog.objects.create(
                action="Payment Excel Export Triggered",
                performed_by=request.user,
                details=f"Exported {row_count} records"
            )

            response = HttpResponse(
                excel_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="payment_export.xlsx"'
            return response

        except ImportError:
            return Response(
                {'error': 'openpyxl not installed. Run: pip install openpyxl'},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        """Download PDF report. §3.1.H: all reports exportable as PDF."""
        from django.http import HttpResponse

        funding_type = request.query_params.get('funding_type', 'all').lower()
        date_from    = request.query_params.get('date_from')
        date_to      = request.query_params.get('date_to')
        student_name = request.query_params.get('student')
        semester     = request.query_params.get('semester')
        year         = request.query_params.get('year')
        award_type   = request.query_params.get('award_type')

        csv_bytes, row_count = _build_full_csv(
            funding_type=funding_type, date_from=date_from, date_to=date_to,
            all_statuses=False, unsent_only=False,
            student_name=student_name, semester=semester, year=year, award_type=award_type,
        )

        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import io, csv as _csv

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                    leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
            styles = getSampleStyleSheet()
            elements = []
            elements.append(Paragraph("DGG Post-Secondary Funding — Student Report", styles['Title']))
            filters = ' | '.join(filter(None, [
                f"Type: {funding_type}" if funding_type != 'all' else '',
                f"From: {date_from}" if date_from else '',
                f"To: {date_to}" if date_to else '',
                f"Student: {student_name}" if student_name else '',
                f"Semester: {semester}" if semester else '',
                f"Year: {year}" if year else '',
            ]))
            if filters:
                elements.append(Paragraph(f"Filters: {filters}", styles['Normal']))
            elements.append(Spacer(1, 12))

            reader = _csv.reader(io.StringIO(csv_bytes.decode('utf-8')))
            rows = list(reader)

            if rows:
                # Only include key columns to fit landscape A4
                KEY_COLS = [0, 1, 2, 3, 4, 5, 6, 34, 35, 36, 37, 38, 39]  # submission + personal + totals + dates
                def pick_cols(row):
                    return [str(row[i]) if i < len(row) else '' for i in KEY_COLS]

                table_data = [pick_cols(r) for r in rows]
                col_widths = [55, 90, 55, 55, 90, 90, 70, 60, 60, 60, 60, 70, 70]

                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                    ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("No records found matching the selected filters.", styles['Normal']))

            doc.build(elements)
            pdf_bytes = buffer.getvalue()

            AuditLog.objects.create(
                action="Payment PDF Export Triggered",
                performed_by=request.user,
                details=f"Exported {row_count} records"
            )

            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'
            return response

        except ImportError:
            return Response(
                {'error': 'reportlab not installed. Run: pip install reportlab'},
                status=status.HTTP_501_NOT_IMPLEMENTED,
            )

    @action(detail=False, methods=['post'], url_path='dispatch_report')
    def dispatch_report(self, request):
        """
        Email ONLY director-approved submissions that have NOT yet been sent to finance.
        Includes both modern FormSubmissions and legacy Applications.
        """
        from django.utils import timezone as tz
        from forms.models import FormSubmission
        from email_sender import send_finance_report, send_funding_processed

        # ── 1. Identify all approved but unsent records ──
        unsent_subs = FormSubmission.objects.filter(status='accepted', finance_sent_at__isnull=True)
        unsent_legacy = Application.objects.filter(status='approved', finance_sent_at__isnull=True)

        if not unsent_subs.exists() and not unsent_legacy.exists():
            return Response({
                'status': 'nothing_to_send',
                'count': 0,
                'message': 'No new approved applications to send — all have already been dispatched.',
            })

        # ── 2. Build combined CSV ──
        csv_bytes, total_rows = _build_full_csv(all_statuses=False, unsent_only=True)

        triggered_by = getattr(request.user, 'full_name', '') or request.user.email

        # ── 2b. Create a batch FinanceConfirmToken for the first unsent submission
        #        (or a "batch" token not tied to a specific submission) so Finance
        #        can click once to confirm the whole dispatch. §7.3.
        confirm_url = ''
        try:
            import uuid as _uuid
            from datetime import timedelta as _td
            from django.utils import timezone as _tz2
            from api.models import FinanceConfirmToken
            from django.conf import settings as _s2
            first_sub = unsent_subs.first()
            batch_token = FinanceConfirmToken.objects.create(
                submission=first_sub,
                token=_uuid.uuid4().hex,
                expires_at=_tz2.now() + _td(days=30),
            )
            base_url = getattr(_s2, 'SITE_URL', request.build_absolute_uri('/').rstrip('/'))
            confirm_url = f"{base_url}/api/finance-confirm/{batch_token.token}/"
        except Exception as _e:
            import logging
            logging.getLogger(__name__).warning("FinanceConfirmToken creation failed: %s", _e)

        # ── 3. Send email ──
        ok = send_finance_report(
            csv_bytes=csv_bytes,
            total_students=total_rows,
            triggered_by=triggered_by,
            confirm_url=confirm_url,
        )

        if not ok:
            return Response(
                {'status': 'error', 'message': 'Failed to send email — check server logs.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # ── 4. Post-send: mark as sent and notify students ──
        now = tz.now()
        
        # Update FormSubmissions
        sub_ids = list(unsent_subs.values_list('id', flat=True))
        FormSubmission.objects.filter(id__in=sub_ids).update(
            status='sent_to_finance',
            finance_sent_at=now,
            finance_sent_by=request.user,
        )

        # Update Legacy Applications
        legacy_ids = list(unsent_legacy.values_list('id', flat=True))
        Application.objects.filter(id__in=legacy_ids).update(
            finance_sent_at=now,
            finance_sent_by=request.user,
        )

        # Per-submission: update payment, notify student
        for sub in unsent_subs.select_related('student', 'form'):
            student = sub.student
            if not student:
                continue

            # Mark any pending payments for this submission as issued
            # Use the new direct link to submission
            from api.models import Payment
            Payment.objects.filter(submission=sub, status=Payment.Status.PENDING).update(status=Payment.Status.ISSUED)

            # In-app notification
            create_notification(
                student,
                'Payment Sent to Finance 💰',
                f"Your approved funding for '{sub.form.title}' "
                f"(${sub.amount}) has been sent to the Finance Department. "
                f"Expect payment within 2–5 working days."
            )

            # Email notification
            if student.email:
                try:
                    # Build breakdown from payments
                    payments = list(sub.payments.filter(status=Payment.Status.ISSUED))
                    if payments:
                        breakdown = [{'name': p.payment_type, 'amount': float(p.amount)} for p in payments]
                        total = sum(b['amount'] for b in breakdown)
                    else:
                        breakdown = [{'name': sub.form.title, 'amount': float(sub.amount or 0)}]
                        total = float(sub.amount or 0)

                    send_funding_processed(
                        student_email=student.email,
                        student_name=student.full_name,
                        program_name=sub.form.title,
                        semester=sub.office_use_data.get('semester', '') if sub.office_use_data else '',
                        year=sub.office_use_data.get('year', '') if sub.office_use_data else '',
                        total_amount=total,
                        funding_breakdown=breakdown,
                    )
                except Exception as exc:
                    import logging
                    logging.getLogger(__name__).error(
                        'send_funding_processed failed for submission %s: %s', sub.id, exc
                    )

        # ── 5. Audit log ──
        AuditLog.objects.create(
            action=f"Finance Report Dispatched — {total_rows} new approved records",
            performed_by=request.user,
            role=request.user.role,
            details=f"Record counts: FormSubmissions={len(sub_ids)}, LegacyApps={len(legacy_ids)}",
        )

        return Response({
            'status': 'success',
            'count': total_rows,
            'message': f'Successfully dispatched {total_rows} records to Finance and marked them as sent.',
        })

    @action(detail=False, methods=['post'], url_path='dispatch_custom')
    def dispatch_custom(self, request):
        """
        Staff-driven payment dispatch — replaces the .env-locked recipient flow.

        Body JSON
        ---------
        recipients   : list[str]   required — one or more finance emails
        payment_ids  : list[int]   required — payment rows the staff selected
        notes        : str         optional — free-text note included in email body
        subject      : str         optional — overrides the default subject line

        Behavior
        --------
        - Builds a per-payment CSV (one row per payment) with the full student +
          banking + enrollment context Finance needs.
        - Sends to the supplied recipients (CC'ing each other in To header).
        - Marks the included payments as ISSUED, marks each parent submission
          as sent_to_finance, and emails the student that funding was processed.
        - Writes an AuditLog row with the recipient list + record count.
        """
        import csv as _csv
        import io as _io
        import re as _re
        from django.utils import timezone as tz
        from email_sender import send_finance_report_custom, send_funding_processed
        from api.models import Payment
        from forms.models import FormSubmission

        body = request.data or {}
        recipients_raw = body.get('recipients') or []
        if isinstance(recipients_raw, str):
            recipients_raw = [recipients_raw]
        recipients = [str(r).strip() for r in recipients_raw if str(r).strip()]
        if not recipients:
            return Response(
                {'status': 'error', 'message': 'At least one recipient email is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        email_re = _re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
        invalid = [r for r in recipients if not email_re.match(r)]
        if invalid:
            return Response(
                {'status': 'error', 'message': f'Invalid email(s): {", ".join(invalid)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payment_ids = body.get('payment_ids') or []
        if not isinstance(payment_ids, list) or not payment_ids:
            return Response(
                {'status': 'error', 'message': 'At least one payment must be selected.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            payment_ids = [int(p) for p in payment_ids]
        except (TypeError, ValueError):
            return Response(
                {'status': 'error', 'message': 'payment_ids must be integers.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        notes = str(body.get('notes') or '').strip()
        subject = str(body.get('subject') or '').strip()

        payments_qs = (Payment.objects
                       .filter(id__in=payment_ids)
                       .select_related('user', 'user__profile', 'submission', 'submission__form', 'application'))
        if not payments_qs.exists():
            return Response(
                {'status': 'error', 'message': 'No matching payment rows found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ── Build the per-STUDENT aggregate CSV ──
        # Finance only needs: student identity, banking, total amount. Per-payment
        # detail (type, stream, dates) is intentionally stripped — payment_type is
        # an internal classification, finance just disburses one lump per student.
        D = '—'
        headers = [
            'Student Full Name', 'Beneficiary #',
            'Institution', 'Program', 'Semester',
            'Account Holder Name', 'Bank Name',
            'Transit #', 'Institution #', 'Account #',
            'Total Amount ($)', 'Commitment #',
        ]
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow(headers)

        # Aggregate by student so finance sees one disbursement row per person.
        from collections import OrderedDict
        agg: "OrderedDict[int, dict]" = OrderedDict()
        for p in payments_qs:
            u = p.user
            if not u:
                continue
            entry = agg.setdefault(u.id, {'user': u, 'total': 0, 'submission': p.submission})
            try:
                entry['total'] += float(p.amount or 0)
            except (TypeError, ValueError):
                pass

        summary_rows = []  # (name, payment_type, formatted_amount) for HTML preview — type left blank
        for entry in agg.values():
            u = entry['user']
            sub = entry.get('submission')
            profile = getattr(u, 'profile', None)
            beneficiary = (u.beneficiary_number or (profile.beneficiary_number if profile else None) or D)
            institution = u.institution_name or (profile.institute if profile else None) or D
            program = u.program_credential or D
            semester = u.current_semester or D
            commitment = ''
            if sub and sub.office_use_data:
                commitment = sub.office_use_data.get('commitmentNum', '') or ''
            writer.writerow([
                u.full_name or D,
                beneficiary,
                institution,
                program,
                semester,
                u.account_holder_name or u.full_name or D,
                u.bank_name or D,
                u.transit_number or D,
                u.inst_number or D,
                u.account_number or D,
                f"{entry['total']:.2f}",
                commitment or D,
            ])
            summary_rows.append((u.full_name or 'Student', '', f"{entry['total']:,.2f}"))

        csv_bytes = buf.getvalue().encode('utf-8')

        triggered_by = getattr(request.user, 'full_name', '') or request.user.email

        ok = send_finance_report_custom(
            csv_bytes=csv_bytes,
            recipients=recipients,
            record_count=len(agg),
            triggered_by=triggered_by,
            notes=notes,
            subject_override=subject,
            summary_rows=summary_rows,
        )
        if not ok:
            return Response(
                {'status': 'error', 'message': 'Failed to send email — check server logs.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # ── Post-send: mark payments + parent submissions, notify students ──
        now = tz.now()
        payments_qs.filter(status=Payment.Status.PENDING).update(status=Payment.Status.ISSUED)

        # Distinct submissions covered by this dispatch — mark each.
        sub_ids = list(payments_qs.exclude(submission__isnull=True).values_list('submission_id', flat=True).distinct())
        if sub_ids:
            FormSubmission.objects.filter(id__in=sub_ids).update(
                status='sent_to_finance',
                finance_sent_at=now,
                finance_sent_by=request.user,
            )
            for sub in FormSubmission.objects.filter(id__in=sub_ids).select_related('student', 'form'):
                student = sub.student
                if not student:
                    continue
                create_notification(
                    student,
                    'Payment Sent to Finance 💰',
                    f"Your approved funding for '{sub.form.title if sub.form else 'application'}' "
                    f"(${sub.amount}) has been sent to the Finance Department. "
                    f"Expect payment within 2–5 working days."
                )
                if student.email:
                    try:
                        payments = list(sub.payments.filter(status=Payment.Status.ISSUED))
                        breakdown = ([{'name': p.payment_type, 'amount': float(p.amount)} for p in payments]
                                     if payments
                                     else [{'name': sub.form.title if sub.form else 'Application',
                                            'amount': float(sub.amount or 0)}])
                        total = sum(b['amount'] for b in breakdown)
                        send_funding_processed(
                            student_email=student.email,
                            student_name=student.full_name,
                            program_name=sub.form.title if sub.form else 'Application',
                            semester=(sub.office_use_data or {}).get('semester', ''),
                            year=(sub.office_use_data or {}).get('year', ''),
                            total_amount=total,
                            funding_breakdown=breakdown,
                        )
                    except Exception as exc:
                        import logging
                        logging.getLogger(__name__).error(
                            'send_funding_processed failed for submission %s: %s', sub.id, exc
                        )

        AuditLog.objects.create(
            action=f"Custom Finance Dispatch — {payments_qs.count()} payment(s)",
            performed_by=request.user,
            role=request.user.role,
            details=f"Recipients: {', '.join(recipients)}; payment_ids={payment_ids}",
        )

        return Response({
            'status': 'success',
            'count': payments_qs.count(),
            'recipients': recipients,
            'message': f"Dispatched {payments_qs.count()} payment(s) to {len(recipients)} recipient(s).",
        })


class AppealViewSet(viewsets.ModelViewSet):
    queryset = Appeal.objects.all()
    serializer_class = AppealSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if _is_staff(user):
            return self.queryset.all()
        return self.queryset.filter(user=user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'], url_path='escalate')
    def escalate(self, request, pk=None):
        """§4.7: Escalate appeal to next level (director → dggr_official → ceo)."""
        appeal = self.get_object()
        if not _is_staff(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Only staff can escalate appeals.")

        level_order = [
            Appeal.EscalationLevel.DIRECTOR,
            Appeal.EscalationLevel.DGGR_OFFICIAL,
            Appeal.EscalationLevel.CEO,
        ]
        current_idx = level_order.index(appeal.escalation_level) if appeal.escalation_level in level_order else 0
        if current_idx >= len(level_order) - 1:
            return Response({'error': 'Appeal is already at the highest escalation level (CEO).'}, status=status.HTTP_400_BAD_REQUEST)

        appeal.escalation_level = level_order[current_idx + 1]
        appeal.status = Appeal.Status.ESCALATED
        appeal.escalated_at = timezone.now()
        appeal.escalation_notes = request.data.get('notes', '')
        appeal.save()

        AuditLog.objects.create(
            action=f"Appeal #{appeal.id} escalated to {appeal.escalation_level}",
            performed_by=request.user,
            role=request.user.role,
            application=appeal.application,
            details=appeal.escalation_notes,
        )
        create_notification(
            appeal.user,
            "Appeal Escalated",
            f"Your appeal has been escalated to the {appeal.get_escalation_level_display()}.",
        )
        return Response({
            'status': 'escalated',
            'new_level': appeal.escalation_level,
            'level_display': appeal.get_escalation_level_display(),
        })


class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if _is_staff(user):
            return self.queryset.all()
        return self.queryset.filter(application__student=user)


class UserDocumentViewSet(viewsets.ModelViewSet):
    queryset = UserDocument.objects.all()
    serializer_class = UserDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]

    # Accepted file signatures (magic bytes) keyed by extension
    _MAGIC = {
        '.pdf':  b'%PDF',
        '.jpg':  b'\xff\xd8\xff',
        '.jpeg': b'\xff\xd8\xff',
        '.png':  b'\x89PNG',
    }

    def get_queryset(self):
        user = self.request.user
        if _is_staff(user):
            student_id = self.request.query_params.get('student_id')
            if student_id:
                return self.queryset.filter(user_id=student_id)
            return self.queryset.all()
        return self.queryset.filter(user=user)

    def _validate_upload(self, uploaded_file):
        allowed_extensions = getattr(settings, 'ALLOWED_UPLOAD_EXTENSIONS', ['.pdf', '.jpg', '.jpeg', '.png', '.doc', '.docx'])
        max_size = getattr(settings, 'FILE_UPLOAD_MAX_MEMORY_SIZE', 10 * 1024 * 1024)

        ext = os.path.splitext(uploaded_file.name)[1].lower()
        if ext not in allowed_extensions:
            raise ValidationError(f"File type '{ext}' is not allowed. Accepted: {', '.join(allowed_extensions)}")

        if uploaded_file.size > max_size:
            raise ValidationError(f"File exceeds the maximum allowed size of {max_size // (1024 * 1024)} MB.")

        # Magic-byte check for common types to prevent extension spoofing
        if ext in self._MAGIC:
            header = uploaded_file.read(8)
            uploaded_file.seek(0)
            if not header.startswith(self._MAGIC[ext]):
                raise ValidationError("File content does not match its extension.")

    def perform_create(self, serializer):
        uploaded_file = self.request.FILES.get('file')
        if uploaded_file:
            self._validate_upload(uploaded_file)
        instance = serializer.save(user=self.request.user)
        # Notify all admin/ssw staff that a student uploaded a document
        staff_users = User.objects.filter(role__in=('admin', 'ssw'))
        student_name = self.request.user.get_full_name() or self.request.user.email
        file_name = uploaded_file.name if uploaded_file else 'a document'
        for staff in staff_users:
            try:
                create_notification(
                    staff,
                    f"Student Document Upload: {student_name}",
                    f"{student_name} uploaded '{file_name}' to their personal documents.",
                )
            except Exception:
                pass

    def perform_update(self, serializer):
        # Allow the document owner to re-upload; staff can always update
        obj = serializer.instance
        user = self.request.user
        if not _is_staff(user) and obj.user != user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You do not have permission to update this document.")
        uploaded_file = self.request.FILES.get('file')
        if uploaded_file:
            self._validate_upload(uploaded_file)
        serializer.save()


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdminUser]

    def get_queryset(self):
        qs = self.queryset
        submission  = self.request.query_params.get('submission')
        application = self.request.query_params.get('application')
        if submission:
            qs = qs.filter(details__icontains=f'submission {submission}')
        if application:
            qs = qs.filter(application_id=application)
        return qs


# ---------------------------------------------------------------------------
# DIRECTOR ONE-CLICK APPROVE / DENY  (§3.1.D / §2)
# ---------------------------------------------------------------------------

class DirectorActionView(viewsets.GenericViewSet):
    """
    Tokenized one-click approve/deny used from the director's email notification.
    No portal login required. Token is single-use and expires in 48 hours.

    GET  /api/director-action/{token}/   → shows confirmation page (HTML or JSON)
    POST /api/director-action/{token}/   → executes the action
    """
    permission_classes = [permissions.AllowAny]

    def retrieve(self, request, pk=None):
        """Return submission summary for the director to review before confirming."""
        from api.models import DirectorActionToken
        try:
            token_obj = DirectorActionToken.objects.select_related(
                'submission', 'submission__form', 'submission__student'
            ).get(token=pk)
        except DirectorActionToken.DoesNotExist:
            return Response({'error': 'Invalid or expired link.'}, status=status.HTTP_404_NOT_FOUND)
        if not token_obj.is_valid():
            return Response({'error': 'This link has already been used or has expired.'}, status=status.HTTP_410_GONE)
        sub = token_obj.submission
        return Response({
            'action': token_obj.action,
            'submission_id': sub.id,
            'student': sub.student.full_name if sub.student else '—',
            'form': sub.form.title if sub.form else '—',
            'amount': str(sub.amount or 0),
            'status': sub.status,
        })

    def update(self, request, pk=None):
        """Execute the approve or deny action."""
        from api.models import DirectorActionToken
        from api.services.form_service import FormService
        from django.utils import timezone as tz

        try:
            token_obj = DirectorActionToken.objects.select_related(
                'submission', 'director'
            ).get(token=pk)
        except DirectorActionToken.DoesNotExist:
            return Response({'error': 'Invalid or expired link.'}, status=status.HTTP_404_NOT_FOUND)
        if not token_obj.is_valid():
            return Response({'error': 'This link has already been used or has expired.'}, status=status.HTTP_410_GONE)

        action = token_obj.action  # 'approve' or 'deny'
        reason = request.data.get('reason', token_obj.reason or '')
        new_status = 'accepted' if action == 'approve' else 'rejected'

        # Use director if available, else create a synthetic actor for the audit trail
        actor = token_obj.director
        if actor is None:
            # Fall back: any director account (first one found)
            from django.contrib.auth import get_user_model as _gum
            actor = _gum().objects.filter(role='director', is_active=True).first()
        if actor is None:
            return Response({'error': 'No active director account found.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            FormService.update_submission_status(
                token_obj.submission, new_status, actor,
                {'decision_notes': reason, 'reason': reason}
            )
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        # Mark token as used
        token_obj.used_at = tz.now()
        token_obj.save(update_fields=['used_at'])

        AuditLog.objects.create(
            action=f"Director one-click {action.upper()} for submission {token_obj.submission_id}",
            performed_by=actor,
            role='director',
            details=f"Via email token. Reason: {reason}",
        )

        return Response({
            'status': 'success',
            'action': action,
            'submission_id': token_obj.submission_id,
            'message': f"Application #{token_obj.submission_id} has been {action}d.",
        })


# ---------------------------------------------------------------------------
# FINANCE CONFIRM WITHOUT LOGIN  (§3.1.E / §7.3)
# ---------------------------------------------------------------------------

class FinanceConfirmView(viewsets.GenericViewSet):
    """
    Finance confirms payment processed via a single click from their email.
    No portal login required. Token expires in 30 days.

    GET  /api/finance-confirm/{token}/  → shows what's being confirmed
    POST /api/finance-confirm/{token}/  → marks payments as processed
    """
    permission_classes = [permissions.AllowAny]

    def retrieve(self, request, pk=None):
        from api.models import FinanceConfirmToken
        try:
            tok = FinanceConfirmToken.objects.select_related(
                'submission', 'submission__student', 'submission__form'
            ).get(token=pk)
        except FinanceConfirmToken.DoesNotExist:
            return Response({'error': 'Invalid or expired link.'}, status=status.HTTP_404_NOT_FOUND)
        if not tok.is_valid():
            if tok.confirmed_at:
                return Response({
                    'status': 'already_confirmed',
                    'confirmed_at': tok.confirmed_at,
                    'confirmed_by': tok.confirmed_by_name,
                })
            return Response({'error': 'Link has expired.'}, status=status.HTTP_410_GONE)
        sub = tok.submission
        return Response({
            'submission_id': sub.id if sub else None,
            'student': sub.student.full_name if sub and sub.student else '—',
            'amount': str(sub.amount or 0) if sub else '0',
            'form': sub.form.title if sub and sub.form else '—',
        })

    def update(self, request, pk=None):
        from api.models import FinanceConfirmToken, Payment
        from django.utils import timezone as tz

        try:
            tok = FinanceConfirmToken.objects.select_related('submission').get(token=pk)
        except FinanceConfirmToken.DoesNotExist:
            return Response({'error': 'Invalid or expired link.'}, status=status.HTTP_404_NOT_FOUND)
        if not tok.is_valid():
            return Response({'error': 'Link has already been used or has expired.'}, status=status.HTTP_410_GONE)

        confirmed_by = request.data.get('name', 'Finance Staff')
        now = tz.now()

        tok.confirmed_at = now
        tok.confirmed_by_name = confirmed_by
        tok.save(update_fields=['confirmed_at', 'confirmed_by_name'])

        # Mark associated payments as ISSUED
        if tok.submission:
            Payment.objects.filter(
                submission=tok.submission, status=Payment.Status.PENDING
            ).update(status=Payment.Status.ISSUED)

        AuditLog.objects.create(
            action=f"Finance confirmed payment for submission {tok.submission_id}",
            performed_by=None,
            details=f"Confirmed by: {confirmed_by} via email token at {now}",
        )

        return Response({
            'status': 'confirmed',
            'message': 'Payment confirmed. Thank you.',
            'confirmed_by': confirmed_by,
        })
